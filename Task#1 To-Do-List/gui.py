from tkinter import *
import tkinter.messagebox as tmsg
from tkinter import ttk
# from PIL import ImageTk, Image
import openpyxl
# from openpyxl import load_workbook
# from openpyxl.workbook import Workbook

# def save_add_task(a,b,c):
#     task_title = title.get().strip()  # Retrieve text and strip leading/trailing whitespace
#     task_detail = details_text.get("1.0", END).strip()  # Retrieve text and strip leading/trailing whitespace
#     task_date = date_entry.get().strip()  # Retrieve text and strip leading/trailing whitespace

#     if not all([task_title, task_detail, task_date]):
#        tmsg.showwarning("Warning", "One or more inputs are empty.")
#        return
#     else:
#        tmsg.showinfo("Task Added","Your task is added successfully!")

def insert_row():
    title_name = title.get()
    description = details_text.get("1.0", END)
    date = date_entry.get()

    #insert Row into excell sheet
    path = "E:\CodSoft Internship\Task#1 To-Do-List\Book (1).xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    if not all([title_name, description, date]):
        tmsg.showwarning("Warning", "One or more inputs are empty.")
        return
    else:
        row_values = [title_name, description,  date]
        sheet.append(row_values)
        workbook.save(path)
        tmsg.showinfo("Task Added","Your task is added successfully!")
        print(title_name,description,date)



def view_task():
    path = "E:\CodSoft Internship\Task#1 To-Do-List\Book.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    print(list_values)


    # title = entry_title.get()
    # description = text_description.get("1.0", "end-1c")
    # due_date = entry_due_date.get()
    
    # listbox_tasks.insert("", "end", values=(title, description, due_date))
    
    # task_title = title.get().strip()  
    # task_detail = details_text.get("1.0", END).strip()  
    # task_date = date_entry.get().strip() 

def task_completed():
    confirmation = tmsg.askquestion("Task completed", "Are you sure you have completed this task?")
    if confirmation :
        tmsg.showinfo("Congratulations!","Congratulations on Completeion of your task!")

#GUI 
root = Tk()
root.title("TO DO LIST APPLICATION")
root.geometry("655x455+300+100")
root.resizable(False, False)
root.config(bg="deep sky blue")

background = "deep sky blue"

title_label = Label(text="Organize Your Tasks, Simplify Your Life", font=('Times New Roman',13,'italic'), padx=200, pady=12)
title_label.pack(side=TOP, fill="x", )

def tab1():

    def add_task():
        #variables
        global title
        global  details_text
        global date_entry
        #destroy tab1 buttons
        create_task.destroy()
        update_task.destroy()
        view_task.destroy()
        deleted_task.destroy()

        # BACK Button FUNCTION
        def back_function():
            title_label.destroy()
            title.destroy()
            detail_label.destroy()
            details_text.destroy()
            date_label.destroy()
            date_entry.destroy()
            back_button.destroy()
            submit_button.destroy()
            root.title("TO DO LIST APPLICATION")  
            tab1()

        # SET WINDOW TITLE
        root.title("ADD TASK")   

        add_task_frame = ttk.LabelFrame(root,text="Add Task")
        add_task_frame.pack()
    

        # Add task buttons and labels
        title_label = Label(root, text="Title", font=('Times New Roman',13), bg=background)
        title_label.place(relx=0.455, rely=0.18)

        title = Entry(root, width=31, font=('arial',13))
        title.place(relx=0.27, rely=0.25)

        detail_label = Label(text="Details",font=('Times New Roman',13), bg=background)
        detail_label.place(relx=0.455, rely=0.33)

        details_text = Text(width=35, height=5)
        details_text.place(relx=0.27, rely=0.4)

        date_label = Label(text="Date",font=('Times New Roman',13), bg=background)
        date_label.place(relx=0.455, rely=0.62)

        date_entry = Entry(root ,width=31,font=('arial',13))
        date_entry.place(relx=0.27, rely=0.69)

        submit_button = Button(text="SAVE", padx=15, pady=5, command=lambda: insert_row())
        submit_button.place(relx=0.33, rely=0.8)

        back_button = Button(text="BACK", padx=17, pady=5, command=back_function)
        back_button.place(relx=0.55, rely=0.8)

    def view_task():    
        #destroy tab1 buttons
        create_task.destroy()
        update_task.destroy()
        view_task.destroy()
        deleted_task.destroy()

        def back_view_task():
            frame_tasks.destroy()
            # columns.destroy()
            listbox_tasks.destroy()
            scrollbar_tasks.destroy()
            back_button_view_tasks.destroy()
            tab1()

        frame_tasks = Frame(root)
        frame_tasks.place(relx=0.02, rely=0.2, relwidth=0.96, relheight=0.6)

        columns = ("Title", "Description", "Due Date")
        listbox_tasks = ttk.Treeview(frame_tasks, columns=columns, show="headings")
        for col in columns:
            listbox_tasks.heading(col, text=col)
            listbox_tasks.column(col, anchor="center")  # Aligning text in the center of each column
        listbox_tasks.pack(side=LEFT, fill=BOTH, expand=True)

        scrollbar_tasks = Scrollbar(frame_tasks, orient=VERTICAL, command=listbox_tasks.yview)
        scrollbar_tasks.pack(side=RIGHT, fill=Y)
        listbox_tasks.config(yscrollcommand=scrollbar_tasks.set)

        #Back button
        back_button_view_tasks = Button(text="BACK", command=back_view_task, padx=20, pady=5)
        back_button_view_tasks.place(relx=0.43, rely=0.85)
        
    def update_task():
         #destroy tab1 buttons
        create_task.destroy()
        update_task.destroy()
        view_task.destroy()
        deleted_task.destroy()

        # Back Function 
        def back_update_task():
            update_task_frame.destroy()
            item_frame.destroy()
            title_label.destroy()
            delete_button.destroy()
            edit_button.destroy()
            back_button_edit_tasks.destroy()
            tab1()

        update_task_frame = Frame(width=550,height=300, bg="black")
        update_task_frame.place(relx=0.08, rely=0.15)

        item_frame = Frame(update_task_frame ,bg="white", width=550 , height=50)
        item_frame.place(relx=0, rely=0)

        title_label=Label(item_frame, text="Fist Task")
        title_label.place(relx=0, rely=0)

        delete_button = Button(item_frame, text="DELETE", padx=8)
        delete_button.place(relx=0.5, rely=0.18)

        edit_button = Button(item_frame, text="EDIT", padx=15)
        edit_button.place(relx=0.628, rely=0.18)

        completed_task_button = Button(item_frame, text="Task Completed", padx=15, command=task_completed)
        completed_task_button.place(relx=0.75, rely=0.18)

        #Back button
        back_button_edit_tasks = Button(text="BACK", command=back_update_task, padx=20, pady=5)
        back_button_edit_tasks.place(relx=0.43, rely=0.85)



    create_task = Button(root, text="Add Task", padx=40, pady=15, command=add_task)
    create_task.place(relx=0.25, rely=0.35)

    update_task = Button(root, text="Update Task", padx=35, pady=15, command=update_task)
    update_task.place(relx=0.55, rely=0.35)

    view_task = Button(root, text="View task List", padx=30, pady=15, command=view_task)
    view_task.place(relx=0.25, rely=0.5)

    deleted_task = Button(root, text="deleted Task", padx=38, pady=15)
    deleted_task.place(relx=0.55, rely=0.5)

tab1()    

root.mainloop()


