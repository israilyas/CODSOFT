from tkinter import *
from tkinter import ttk
import ttkbootstrap as tb
from tkcalendar import *
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

combo_list = ["Subscribed","Not subscribed","Other"]

def load_data():
    path = "Book (1).xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Extract column names from the first row
    header_row = next(sheet.iter_rows(values_only=True))
    column_names = header_row[0]

    # Print column names and indices for debugging purposes
    for index, col_name in enumerate(column_names):
        print(f"Column {index}: {col_name}")

    # Check if the number of columns in Treeview matches the number of columns in Excel file
    # if len(cols) != len(column_names):
    #     print("Number of columns in Treeview doesn't match number of columns in Excel file.")
    #     return

    # Iterate over the column names and add them as headings to the Treeview
    for col_index, col_name in enumerate(column_names):
        if col_name is not None:  # Ensure column name is not None
            print(f"Adding heading for column {col_index}: {col_name}")
            treeview.heading(cols, text=col_name)  # Use column identifier from cols tuple



# def load_data():
#     path = "E:/CodSoft Internship/Task#1 To-Do-List/Book (1).xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook.active

#     # Extract column names from the first row
#     header_row = next(sheet.iter_rows(values_only=True))
#     column_names = header_row[0]

#     # Print column names and indices for debugging purposes
#     for index, col_name in enumerate(column_names):
#         print(f"Column {index}: {col_name}")

#     # Check if the number of columns in Treeview matches the number of columns in Excel file
#     if len(cols) != len(column_names):
#         print("Number of columns in Treeview doesn't match number of columns in Excel file.")
#         return

#     # Iterate over the column names and add them as headings to the Treeview
#     for col_index, col_name in enumerate(column_names):
#         if col_name is not None:  # Ensure column name is not None
#             print(f"Adding heading for column {col_index}: {col_name}")
#             treeview.heading(cols[col_index], text=col_name)  # Use column name from Treeview definition


# def data():
#     #creat workbook instance
#     wb = Workbook()
#     wb = load_workbook("Book (1).xlsx")
#     # active worksheet
#     ws = wb.active

#     list_values = list(ws.values)
#     print(list_values)


# def load_data():
#     path = "E:/CodSoft Internship/Task#1 To-Do-List/Book (1).xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook.active

#     list_values = list(sheet.values)
#     print(list_values)
#     # for col_name in list_values[0]:
#     #     treeview.heading(col_name,text=col_name)
#      # Extract column names from the first row
#     header_row = next(sheet.iter_rows(values_only=True))
#     column_names = header_row[0]

#     # Iterate over the column names and add them as headings to the Treeview
#     for col_name in column_names:
#         if col_name is not None:  # Ensure column name is not None
#             treeview.heading(col_name, text=col_name)

# def load_data():
#     path = "E:/CodSoft Internship/Task#1 To-Do-List/Book (1).xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook.active

#     # Extract column names from the first row
#     header_row = next(sheet.iter_rows(values_only=True))
#     column_names = header_row[0]

#     # Print column names and indices for debugging purposes
#     for index, col_name in enumerate(column_names):
#         print(f"Column {index}: {col_name}")

#     # Iterate over the column names and add them as headings to the Treeview
#     for col_index, col_name in enumerate(column_names):
#         if col_name is not None:  # Ensure column name is not None
#             print(f"Adding heading for column {col_index}: {col_name}")
#             treeview.heading(col_index, text=col_name)


root = Tk()
# root.geometry("500x500")
style = ttk.Style(root)
# root.tk.call("source","forest-dark.tcl")
# root.tk.call("source","forest-light.tcl")
# style.theme_use("forest-dark")


frame = ttk.Frame(root)
frame.pack()

widget_frame = ttk.LabelFrame(frame,text="Add Task")
widget_frame.grid(row=0,column=0, padx=20, pady=10)

title_entry = ttk.Entry(widget_frame )
title_entry.insert("0","Title")
title_entry.bind("<FocusIn>", lambda dlt: title_entry.delete("0",END))
title_entry.grid(row=0, column=0, sticky="ew", padx=5, pady=(0,5))

date_entry = tb.DateEntry(widget_frame, bootstyle="success")
date_entry.grid(row=1, column=0, sticky="ew", padx=5, pady=(0,5))

age_spinbox = ttk.Spinbox(widget_frame, from_=10, to=100)
age_spinbox.insert("0","Age")
age_spinbox.bind("<FocusIn>", lambda dlt: age_spinbox.delete("0",END))
age_spinbox.grid(row=2, column=0, sticky="ew", padx=5, pady=(0,5))

status_combobox = ttk.Combobox(widget_frame, values=combo_list)
status_combobox.current(0)
status_combobox.grid(row=3, column=0, sticky="ew", padx=5, pady=(0,5))

a = BooleanVar()
checkbutton = ttk.Checkbutton(widget_frame, text="Employed", variable=a)
checkbutton.grid(row=4, column=0, sticky="nsew", padx=5, pady=(0,5))

button = ttk.Button(widget_frame, text="Insert")
button.grid(row=5, column=0, sticky="nsew", padx=5, pady=(0,5))

separator = ttk.Separator(widget_frame)
separator.grid(row=6, column=0, padx=(20,10), pady=10, sticky="ew")

mode_switch = ttk.Checkbutton(widget_frame, text="Mode" , style="Switch.TButton")
mode_switch.grid(row=7,column=0, padx=5, pady=10, sticky="nsew")


# TREEVIEW
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side=RIGHT, fill=Y)

cols =("Name","Age","Subscription","Emplayement")
treeview = ttk.Treeview(treeFrame, show="headings", columns=cols, height=13, yscrollcommand=treeScroll.set)

treeview.column("Name", width=100)
treeview.column("Age", width=50)
treeview.column("Subscription", width=100)
treeview.column("Emplayement", width=100)

# Create Headings
# treeview.heading("Name",text="Name")
# treeview.heading("Age",text="Age", anchor=W)
# treeview.heading("Subscription",text="Subscription", anchor=CENTER)
# treeview.heading("Emplayement",text="Emplayement", anchor=W)


treeview.pack()
treeScroll.config(command=treeview.yview)

load_data()


root.mainloop()

